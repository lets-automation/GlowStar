"""Test the Excel report generator produces a well-formed, honest report."""

from __future__ import annotations

import openpyxl

from glowstar.reporting.excel_report import build


def test_report_builds_with_expected_sheets(tmp_path):
    out = build(n=15, out=tmp_path / "report.xlsx")
    wb = openpyxl.load_workbook(out)
    assert set(wb.sheetnames) == {
        "Overview", "Pricing Results", "Accuracy Summary", "Market Research", "Legend & Honesty"
    }
    results = wb["Pricing Results"]
    assert results.max_row == 16          # 15 stones + header
    header = [c.value for c in results[1]]
    # The comparison columns the client asked for must be present.
    for col in ("ACTUAL Disc%", "SUGGESTED Disc%", "Disc Error(pts)",
                "Market median Disc%", "Why (explanation)", "BGM state"):
        assert col in header


def test_parse_stone_list_layout(tmp_path):
    """The simple LAB/Lot No/Shape/Weight/... layout parses into the engine schema."""
    import pandas as pd
    from glowstar.reporting.price_file import parse_stone_file
    df = pd.DataFrame([
        {"#": 1, "LAB": "GIA", "Lot No": "7551532052", "Shape": "ROUND", "Weight": 0.31,
         "Color": "F", "Clarity": "VS2", "Cut": "EX", "Polish": "EX", "Sym": "EX", "Fluor": "NON"},
        {"#": 2, "LAB": "GIA", "Lot No": "6551519792", "Shape": "ROUND", "Weight": 0.40,
         "Color": "G", "Clarity": "VS1", "Cut": "VG", "Polish": "EX", "Sym": "VG", "Fluor": "FNT"},
    ])
    p = tmp_path / "list.xlsx"
    df.to_excel(p, index=False)
    out = parse_stone_file(p)
    assert len(out) == 2
    assert list(out["CPS"]) == ["3EX", "VG"]          # cut leads; 3EX only if all EX
    assert list(out["Fluorescence"]) == ["Non", "Fnt"]
    assert out["Rap"].notna().all()                    # Rap looked up for both
    assert list(out["Shape_full"]) == ["Round", "Round"]
